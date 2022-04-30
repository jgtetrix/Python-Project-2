from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random

wb = Workbook()

# Create destination file
destFile = 'Grades.xlsx'

# Assign the Active Worksheet
ws1 = wb.active

# Assign title to Active Worksheet
ws1.title = "Grades"

# Insert random grades into cells from 1 to 25
for row in range(1, 25):
	grade = random.randrange(60, 90)
	ws1.append([grade])

# Save the excel file
wb.save(filename = destFile)

# Additional Usages:

# Reading an existing workbook
# from openpyxl import load_workbook
# wb = load_workbook(filename = 'empty_book.xlsx')

# Inserting an image
# from openyxl import Workbook
# from openpyxl.drawing.image import Image
